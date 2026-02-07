# moc/core/postprocess/report_crossings.py
from __future__ import annotations

from dataclasses import dataclass
from typing import List, Dict

import numpy as np
import pandas as pd

from moc.core.postprocess.crossings import CrossingResult
from moc.core.postprocess.geometry import GeometryProfile


# -------------------------
# Data model
# -------------------------

@dataclass(frozen=True)
class CrossingEngineeringRow:
    cruce_id: str
    nombre: str
    s_m: float

    z_m: float

    hmin_m: float
    hmax_m: float

    # Presiones manométricas:
    pmin_mca: float            # (Hmin - z)
    pmax_mca: float            # (Hmax - z)

    # Conversión a bar (manométrico)
    pmin_bar: float
    pmax_bar: float

    # Cavitación (margen vs hvap)
    hvap_m: float
    margen_cavitacion_mca: float  # pmin_mca - hvap_m
    cavita: bool

    # Flags “gobernantes”
    is_overpressure: bool      # top_n por pmax_mca (mayor)
    is_underpressure: bool     # top_n por pmin_mca (menor)
    is_cavitation: bool        # top_n por margen (menor) dentro de cavitantes

    # Resumen legible
    criterio_gobierno: str     # "sobrepresion+depresion" etc. o "ninguno"


# -------------------------
# Helpers
# -------------------------

def _interp_linear(x: np.ndarray, y: np.ndarray, xq: float) -> float:
    """Interpolación lineal escalar y(xq) con x monótono."""
    if xq < float(x[0]) or xq > float(x[-1]):
        raise ValueError(
            f"El punto s_m={xq} está fuera del rango del perfil [{float(x[0])}, {float(x[-1])}] m."
        )

    j = int(np.searchsorted(x, xq, side="right"))
    j0 = max(0, j - 1)
    j1 = min(len(x) - 1, j)

    x0, x1 = float(x[j0]), float(x[j1])
    y0, y1 = float(y[j0]), float(y[j1])

    if x1 == x0:
        return y0

    w = (xq - x0) / (x1 - x0)
    return y0 * (1.0 - w) + y1 * w


def mca_to_bar(h_mca: float, *, rho_kg_m3: float = 1000.0, g_m_s2: float = 9.80665) -> float:
    """Convierte mca (≈ m de columna de agua) a bar (manométrico)."""
    return (rho_kg_m3 * g_m_s2 * float(h_mca)) / 1e5


def _make_criterio(*, over: bool, under: bool, cav: bool) -> str:
    flags: List[str] = []
    if over:
        flags.append("sobrepresion")
    if under:
        flags.append("depresion")
    if cav:
        flags.append("cavitacion")
    return "+".join(flags) if flags else "ninguno"


# -------------------------
# Public API
# -------------------------

def crossings_engineering_report(
    crossings: List[CrossingResult],
    geom: GeometryProfile,
    *,
    hvap_m: float,
    rho_kg_m3: float = 1000.0,
    g_m_s2: float = 9.80665,
    top_n: int = 1,
) -> List[CrossingEngineeringRow]:
    """
    Reporte de ingeniería por cruce usando:
      - z(s) interpolado
      - p_mca = H - z
      - margen_cavitacion = pmin_mca - hvap_m

    Flags:
      - is_overpressure: top_n por pmax_mca (mayor)
      - is_underpressure: top_n por pmin_mca (menor)
      - is_cavitation: top_n por margen (menor) dentro de cavitantes
    """
    if not crossings:
        return []

    n = max(1, int(top_n))

    # Pre-cálculo
    tmp = []
    for c in crossings:
        zc = _interp_linear(geom.s_m, geom.z_m, float(c.s_m))

        pmin_mca = float(c.hmin_m) - float(zc)
        pmax_mca = float(c.hmax_m) - float(zc)

        margen = pmin_mca - float(hvap_m)
        cavita = bool(getattr(c, "cav_any", False)) or (margen < 0.0)

        tmp.append(
            {
                "cruce": c,
                "z": float(zc),
                "pmin_mca": float(pmin_mca),
                "pmax_mca": float(pmax_mca),
                "margen": float(margen),
                "cavita": bool(cavita),
            }
        )

    # Top-N sets (por ID)
    over_set = {t["cruce"].crossing_id for t in sorted(tmp, key=lambda r: r["pmax_mca"], reverse=True)[:n]}
    under_set = {t["cruce"].crossing_id for t in sorted(tmp, key=lambda r: r["pmin_mca"])[:n]}

    cav_tmp = [t for t in tmp if t["cavita"]]
    cav_set = set()
    if cav_tmp:
        cav_set = {t["cruce"].crossing_id for t in sorted(cav_tmp, key=lambda r: r["margen"])[:n]}

    rows: List[CrossingEngineeringRow] = []
    for t in tmp:
        c = t["cruce"]
        cid = c.crossing_id

        is_over = cid in over_set
        is_under = cid in under_set
        is_cav = cid in cav_set

        pmin_mca = float(t["pmin_mca"])
        pmax_mca = float(t["pmax_mca"])

        rows.append(
            CrossingEngineeringRow(
                cruce_id=str(cid),
                nombre=str(c.name),
                s_m=float(c.s_m),

                z_m=float(t["z"]),

                hmin_m=float(c.hmin_m),
                hmax_m=float(c.hmax_m),

                pmin_mca=pmin_mca,
                pmax_mca=pmax_mca,

                pmin_bar=float(mca_to_bar(pmin_mca, rho_kg_m3=rho_kg_m3, g_m_s2=g_m_s2)),
                pmax_bar=float(mca_to_bar(pmax_mca, rho_kg_m3=rho_kg_m3, g_m_s2=g_m_s2)),

                hvap_m=float(hvap_m),
                margen_cavitacion_mca=float(t["margen"]),
                cavita=bool(t["cavita"]),

                is_overpressure=bool(is_over),
                is_underpressure=bool(is_under),
                is_cavitation=bool(is_cav),

                criterio_gobierno=_make_criterio(over=is_over, under=is_under, cav=is_cav),
            )
        )

    return rows


# -------------------------
# Exports
# -------------------------

def export_crossings_engineering_csv(rows: List[CrossingEngineeringRow], path_csv: str) -> None:
    df = pd.DataFrame([r.__dict__ for r in rows])
    df.to_csv(path_csv, index=False)


def export_crossings_engineering_excel(
    rows: List[CrossingEngineeringRow],
    path_xlsx: str,
    *,
    sheet_name: str = "Cruces",
) -> None:
    """
    Exporta un Excel con 2 filas de encabezado:
      - fila 1: variable
      - fila 2: unidad
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    header_font = Font(name="Verdana", bold=True)
    units_font = Font(name="Verdana", bold=False)
    center = Alignment(horizontal="center", vertical="center")

    cols = [
        "cruce_id", "nombre", "s_m",
        "z_m",
        "hmin_m", "hmax_m",
        "pmin_mca", "pmax_mca",
        "pmin_bar", "pmax_bar",
        "hvap_m", "margen_cavitacion_mca",
        "cavita",
        "is_overpressure", "is_underpressure", "is_cavitation",
        "criterio_gobierno",
    ]

    units: Dict[str, str] = {
        "cruce_id": "",
        "nombre": "",
        "s_m": "m",
        "z_m": "m",
        "hmin_m": "m",
        "hmax_m": "m",
        "pmin_mca": "mca",
        "pmax_mca": "mca",
        "pmin_bar": "bar",
        "pmax_bar": "bar",
        "hvap_m": "mca",
        "margen_cavitacion_mca": "mca",
        "cavita": "",
        "is_overpressure": "",
        "is_underpressure": "",
        "is_cavitation": "",
        "criterio_gobierno": "",
    }

    # Header row 1
    for j, c in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=j, value=c)
        cell.font = header_font
        cell.alignment = center

    # Units row 2
    for j, c in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=j, value=units.get(c, ""))
        cell.font = units_font
        cell.alignment = center

    # Data from row 3
    for i, r in enumerate(rows, start=3):
        d = r.__dict__
        for j, c in enumerate(cols, start=1):
            ws.cell(row=i, column=j, value=d.get(c))

    ws.freeze_panes = "A3"

    # Column widths
    for j, c in enumerate(cols, start=1):
        letter = get_column_letter(j)
        ws.column_dimensions[letter].width = max(10, min(22, len(c) + 6))

    wb.save(path_xlsx)
