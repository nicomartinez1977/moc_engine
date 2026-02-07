from __future__ import annotations
# moc/core/postprocess/timeseries.py
from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional

import numpy as np
import pandas as pd

from moc.core.models.network import Network
from moc.core.postprocess.summary import ProfileSummary
from moc.core.postprocess.geometry import GeometryProfile
from moc.core.solver.moc_solver import MocResults


# =========================
# Helpers (interpolación)
# =========================

def _interp_spatial_scalar(x: np.ndarray, y: np.ndarray, xq: float) -> float:
    """Interpolación lineal escalar y(xq) con x monótono (estricto)."""
    if xq < float(x[0]) or xq > float(x[-1]):
        raise ValueError(
            f"El punto s_m={xq} está fuera del rango del perfil [{float(x[0])}, {float(x[-1])}] m."
        )

    j = int(np.searchsorted(x, xq, side="right"))
    j0 = max(0, j - 1)
    j1 = min(len(x) - 1, j)

    x0, x1 = float(x[j0]), float(x[j1])
    y0, y1 = float(y[j0]), float(y[j1])

    if x1 == x0:
        return y0

    w = (xq - x0) / (x1 - x0)
    return y0 * (1.0 - w) + y1 * w


def _interp_series_spatial(res: MocResults, s: np.ndarray, s_q: float) -> np.ndarray:
    """Interpola H(t, s_q) espacialmente para todos los tiempos."""
    if s_q < float(s[0]) or s_q > float(s[-1]):
        raise ValueError(f"El punto s_m={s_q} está fuera del rango [{float(s[0])}, {float(s[-1])}] m.")

    j = int(np.searchsorted(s, s_q, side="right"))
    j0 = max(0, j - 1)
    j1 = min(len(s) - 1, j)

    s0, s1 = float(s[j0]), float(s[j1])
    if s1 == s0:
        return res.H[:, j0].copy()

    w = (s_q - s0) / (s1 - s0)
    return (1.0 - w) * res.H[:, j0] + w * res.H[:, j1]


def downsample_by_time(df: pd.DataFrame, *, write_every_s: float, t_col: str = "t") -> pd.DataFrame:
    """
    Downsample por tiempo (mantiene primer y último punto).
    - write_every_s <= 0 => devuelve df completo
    Requiere que df[t_col] sea creciente.
    """
    if write_every_s is None or float(write_every_s) <= 0:
        return df

    dt = float(write_every_s)
    t = df[t_col].to_numpy(dtype=float)

    if t.size == 0:
        return df

    keep = np.zeros_like(t, dtype=bool)
    keep[0] = True
    last = t[0]

    for i in range(1, len(t) - 1):
        if t[i] - last >= dt:
            keep[i] = True
            last = t[i]

    keep[-1] = True
    return df.loc[keep].reset_index(drop=True)


# =========================
# Public API
# =========================

@dataclass(frozen=True)
class CrossingSpec:
    cruce_id: str
    nombre: str
    s_m: float


def read_crossings_csv(path_csv: str) -> List[CrossingSpec]:
    df = pd.read_csv(path_csv)
    required = {"cruce_id", "nombre", "s_m"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"El archivo de cruces no contiene las columnas requeridas: {sorted(missing)}")

    out: List[CrossingSpec] = []
    for _, r in df.iterrows():
        out.append(
            CrossingSpec(
                cruce_id=str(r["cruce_id"]).strip(),
                nombre=str(r["nombre"]).strip(),
                s_m=float(r["s_m"]),
            )
        )
    return out


def extract_timeseries_at_s_mca(
    network: Network,
    res: MocResults,
    prof: ProfileSummary,
    geom: GeometryProfile,
    *,
    s_m: float,
) -> pd.DataFrame:
    """
    Serie tiempo en posición s_m:
      columnas: t, H, z, P

    Unidades:
      t [s]
      H [m]   (cota piezométrica)
      z [m]   (cota geométrica)
      P [mca] (presión manométrica como H - z)

    Nota: aquí no usamos Pa/bar; todo en m / mca.
    """
    # H(t)
    H_ts = _interp_series_spatial(res, prof.s_m, float(s_m))

    # z(s)
    z_q = _interp_spatial_scalar(geom.s_m, geom.z_m, float(s_m))

    # Presión manométrica en mca
    P_ts = H_ts - z_q

    return pd.DataFrame(
        {
            "t": np.asarray(res.times, dtype=float),
            "H": np.asarray(H_ts, dtype=float),
            "z": np.full_like(H_ts, z_q, dtype=float),
            "P": np.asarray(P_ts, dtype=float),
        }
    )


def extract_timeseries_for_crossings_mca(
    network: Network,
    res: MocResults,
    prof: ProfileSummary,
    geom: GeometryProfile,
    crossings: List[CrossingSpec],
) -> Dict[str, pd.DataFrame]:
    """
    Retorna dict: cruce_id -> DataFrame (t, H, z, P).
    """
    out: Dict[str, pd.DataFrame] = {}
    for c in crossings:
        out[c.cruce_id] = extract_timeseries_at_s_mca(network, res, prof, geom, s_m=c.s_m)
    return out


# =========================
# Exports
# =========================

_UNITS = {"t": "s", "H": "m", "z": "m", "P": "mca"}


def export_timeseries_to_csv_folder(series_by_id: Dict[str, pd.DataFrame], folder: str) -> List[str]:
    """
    Exporta CSV por cada serie (sin fila de unidades).
    Retorna lista de paths generados.
    """
    import os

    os.makedirs(folder, exist_ok=True)
    out_paths: List[str] = []
    for cid, df in series_by_id.items():
        path = os.path.join(folder, f"serie_{cid}.csv")
        df.to_csv(path, index=False)
        out_paths.append(path)
    return out_paths


def export_timeseries_to_excel(
    series_by_id: Dict[str, pd.DataFrame],
    out_xlsx: str,
    *,
    write_every_s: float = 0.0,
    name_by_id: Optional[Dict[str, str]] = None,
    y_cols_for_chart: Iterable[str] = ("H",),
) -> None:
    """
    Exporta un Excel multi-hoja:
    - Fila 1: nombres de variables (negrita, centrado)
    - Fila 2: unidades (centrado)
    - Datos desde fila 3
    - Gráfico Scatter Smooth (H vs t por defecto) con:
        - título = nombre de hoja
        - ejes: "t [s]" y "<var> [unidad]"
    - write_every_s: downsample por tiempo para aligerar el archivo.

    name_by_id:
      si se entrega, el nombre de hoja será f"{cruce_id} - {nombre}" (truncado a 31 chars).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.chart import ScatterChart, Reference, Series
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    # elimina hoja por defecto
    wb.remove(wb.active)

    header_font = Font(name="Verdana", bold=True)
    units_font = Font(name="Verdana", bold=False)
    header_align = Alignment(horizontal="center", vertical="center")
    units_align = Alignment(horizontal="center", vertical="center")

    for cid, df0 in series_by_id.items():
        df = downsample_by_time(df0, write_every_s=float(write_every_s), t_col="t")

        # nombre de hoja (máx 31)
        sheet_name = cid
        if name_by_id and cid in name_by_id and name_by_id[cid]:
            sheet_name = f"{name_by_id[cid]}"
        sheet_name = sheet_name[:31]

        ws = wb.create_sheet(title=sheet_name)

        # orden de columnas: t, H, z, P (si existen)
        preferred = [c for c in ["t", "H", "z", "P"] if c in df.columns]
        other = [c for c in df.columns if c not in preferred]
        cols = preferred + other

        # --- escribir encabezados
        for j, col in enumerate(cols, start=1):
            c1 = ws.cell(row=1, column=j, value=col)
            c1.font = header_font
            c1.alignment = header_align

            unit = _UNITS.get(col, "")
            c2 = ws.cell(row=2, column=j, value=unit)
            c2.font = units_font
            c2.alignment = units_align

        # --- datos
        for i in range(len(df)):
            for j, col in enumerate(cols, start=1):
                ws.cell(row=3 + i, column=j, value=float(df.iloc[i][col]))

        # --- formato columnas (ancho)
        for j, col in enumerate(cols, start=1):
            letter = get_column_letter(j)
            ws.column_dimensions[letter].width = max(10, min(18, len(col) + 6))

        # congelar panes debajo de encabezados
        ws.freeze_panes = "A3"

        # --- gráfico H vs t (o y_cols_for_chart)
        y_cols = [c for c in y_cols_for_chart if c in cols]
        if ("t" in cols) and y_cols and len(df) > 1:
            chart = ScatterChart()
            chart.smooth = True
            chart.title = sheet_name
            chart.x_axis.title = "t [s]"

            if len(y_cols) == 1:
                y0 = y_cols[0]
                chart.y_axis.title = f"{y0} [{_UNITS.get(y0, '')}]".rstrip()
            else:
                chart.y_axis.title = " / ".join([f"{y} [{_UNITS.get(y,'')}]" for y in y_cols]).rstrip()

            x_col_idx = cols.index("t") + 1
            xvalues = Reference(ws, min_col=x_col_idx, min_row=3, max_row=2 + len(df))

            for y in y_cols:
                y_col_idx = cols.index(y) + 1
                values = Reference(ws, min_col=y_col_idx, min_row=3, max_row=2 + len(df))
                series = Series(values, xvalues, title=y)
                chart.series.append(series)

            # posición del gráfico
            ws.add_chart(chart, "G3")

    wb.save(out_xlsx)
